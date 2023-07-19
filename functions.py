# Importar todas las bibliotecas necesarias
import pandas as pd
import tkinter as tk
from tkinter import filedialog
import os
import requests
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side
import tempfile
import logging

def get_file_path():
    input('Presione Enter para seleccionar un archivo para leer en un dataframe...')
    root = tk.Tk()
    root.withdraw()
    file_path = filedialog.askopenfilename()

    # Asegurarse de que el archivo seleccionado sea un archivo de Excel que pueda ser procesado
    if not file_path.endswith('.xlsx'):
        print('Por favor, elija un archivo de Excel.')
        exit()

    # Crear una nueva carpeta llamada "UF_modelo_{nombre de archivo original}_{fecha actual}"
    new_folder_name = f'UF_modelo_{os.path.basename(file_path)}_{datetime.now().strftime("%Y-%d-%m_%H-%M-%S")}'
    if not os.path.exists(new_folder_name):
        os.mkdir(new_folder_name)

    # Cambiar el directorio de trabajo actual a la nueva carpeta
    os.chdir(new_folder_name)

    # Configuración del logging
    log_file_name = f'log_{os.path.basename(file_path)}_{datetime.now().strftime("%Y-%d-%m_%H-%M-%S")}.log'
    logging.basicConfig(filename=log_file_name, level=logging.INFO, 
                        format='%(asctime)s %(levelname)s %(message)s', 
                        datefmt='%m/%d/%Y %I:%M:%S %p',
                        )
    
    logging.info('Inició la función get_file_path')

    # Mensaje de diálogo confirmando que el archivo ha sido seleccionado y leído
    print(f'El archivo {os.path.basename(file_path)} ha sido leído en un dataframe.')
    logging.info(f'El archivo {os.path.basename(file_path)} ha sido leído en un dataframe.')

    return file_path

def load_dataframe(file_path):
    logging.info(f'Cargando dataframe desde {file_path}')
    return pd.read_excel(file_path)

def find_code_column(df):
    code_column = None
    code_prefixes = [(4, 'REV'), (5, 'RE'), (6, 'NRE'), (7, 'C&GE'), (8, 'OI&E'), (9, 'OI&E')]
    groups = {i: {"codes": [], "names": [], "final": [], "prefix": prefix} for i, prefix in code_prefixes}
    logging.info(f'Se han creado los prefijos de código basados en el primer dígito del código: {code_prefixes}.')
    logging.info(f'Se han creado grupos basados en los prefijos de código: {groups}.')

    # Encontrar la columna que contiene los códigos basándose en el código de ingresos
    for col in df.columns:
        if '40100-00' in df[col].values:
            code_column = col
            break
    logging.info(f'Se ha encontrado la columna de código: {code_column}')
    return code_column, code_prefixes, groups

def group_data_by_code(df, code_column, code_prefixes, groups):
    if code_column is None:
        print("No se pudo encontrar la columna de ingresos.")
        logging.info('No se pudo encontrar la columna de ingresos.')
    else:
        name_column = df.columns[df.columns.get_loc(code_column) + 1]
        logging.info(f'Se ha creado la columna de nombres basada en la posición de la columna de códigos.')

        # Para cada grupo, filtrar las filas basándose en el código, crear un diccionario y poblar las listas
        for i, prefix in code_prefixes:
            group_rows = df[(df[code_column] >= f'{i}0000-00') & (df[code_column] <= f'{i}9999-99')]
            group_data = {(code, f"{prefix} - " + group_rows.loc[group_rows[code_column] == code, name_column].values[0].strip().title()): [] for code in group_rows[code_column]}  

                
            for (code, name), data in group_data.items():
                groups[i]["codes"].append(code)
                groups[i]["names"].append(name)
                groups[i]["final"].append(code + ": " + name)

        logging.info('Se han poblado los grupos para el grupo basado en el prefijo del código.')
        logging.info('Se han creado las filas de grupo para el grupo basado en el prefijo del código.')
        logging.info('Se han creado los datos del grupo para el grupo basado en el prefijo del código.')

    return groups

def populate_data_dict(df, code_column, codes, dates):
    data_dict = {}
    # Iterar sobre cada columna
    for col in df.columns:
        # Iterar sobre cada fila en la columna
        for cell in df[col]:
            try:
                # Intentar convertir la celda a un datetime
                date = pd.to_datetime(cell, format='%b %Y', errors='coerce')

                # Si la celda es una fecha
                if pd.notnull(date):
                    date_str = cell
                    dates.append(date_str)
                    data_dict[date_str] = {}
                    

                    for code in codes:
                        # Obtener la fila donde el código coincide en el DataFrame original
                        row = df[df[code_column] == code]

                        # Si hay alguna fila coincidente, obtener el valor correspondiente de la fila con el código
                        if not row.empty:
                            value = row.iat[0, df.columns.get_loc(col)]
                            data_dict[date_str][code] = value

                    # Salir del bucle interno una vez que se encuentra la primera fecha en la columna
                    break
            except ValueError:
                # Si la celda no pudo ser convertida a una fecha, continuar a la siguiente celda
                continue
    return data_dict, dates

def fetch_exchange_rate(dates):
    try:
        # si el servidor no está caído
        url = 'https://www.alphavantage.co/query?function=FX_MONTHLY&from_symbol=CLF&to_symbol=CLP&apikey=5TLIDN7TN6IQZJUE'
        r = requests.get(url)
        data = r.json()

        # Si hay un mensaje de error en la respuesta, devolver divisores por defecto
        if 'Error Message' in data:
            print("El servidor está caído. Usando divisores por defecto.")
            logging.info('El servidor está caído. Usando divisores por defecto.')
            return [1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1]

        # Extraer los datos de la tasa de cambio mensual de la respuesta
        monthly_rates = data['Time Series FX (Monthly)']

        # Lista para almacenar las tasas de cambio
        exchange_rates = []

        # Iterar sobre los datos mensuales y extraer las tasas de cambio
        for date, rate in monthly_rates.items():
            # Convertir la cadena de la fecha a un objeto datetime
            date_obj = datetime.strptime(date, '%Y-%m-%d')
            # Formatear la fecha como "Mes Año" (por ejemplo, "Jun 2022")
            month_year = date_obj.strftime('%b %Y')

            # Comprobar si el mes/año está en la lista de fechas
            if month_year in dates:
                exchange_rate = rate['4. close']
                exchange_rates.append((month_year, exchange_rate))

        # Ordenar las tasas de cambio basándose en el mes/año en orden ascendente
        exchange_rates.sort(key=lambda x: datetime.strptime(x[0], '%b %Y'))

        # Extraer las tasas de cambio de la lista de tuplas
        divisors_UF = [x[1] for x in exchange_rates]

        # Convertir las tasas de cambio a flotantes
        divisors_UF = [float(x) for x in divisors_UF]

        print("Se ha obtenido con éxito los datos del servidor.")
        logging.info('Se ha obtenido con éxito los datos del servidor.')
        return divisors_UF

    except Exception as e:
        print(f'Error al obtener la tasa de cambio: {e}')
        print("Usando divisores por defecto debido al error.")
        logging.info(f'Error al obtener la tasa de cambio: {e}')
        logging.info('Usando divisores por defecto debido al error.')
        # En caso de cualquier otra excepción no manejada, devolver divisores por defecto
        return [1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1]

def create_and_format_column(df, new_col_name, codes, divisors=None, rounding=0, position=None, factor=1., subtract_codes=None):
    subtract_cols = []
    cols_to_sum = [col for col in df.columns if any(code in col for code in codes)]
    logging.info(f'Columnas para ser sumadas para {new_col_name}: {cols_to_sum}') 

    if subtract_codes is not None:
        subtract_cols = [col for col in df.columns if any(code in col for code in subtract_codes)]
        for col in subtract_cols:
            if col in cols_to_sum:  # comprobar si la columna está en cols_to_sum antes de remover
                cols_to_sum.remove(col)  

    logging.info(f'Columnas para ser restadas para {new_col_name}: {subtract_cols}')
    # print(f'Columnas para ser sumadas para {new_col_name}: {cols_to_sum}')  # Línea de depuración
    # print(f'Columnas para ser restadas para {new_col_name}: {subtract_cols}')  # Línea de depuración

    df[new_col_name] = df[cols_to_sum].sum(axis=1)
    df[new_col_name] -= df[subtract_cols].sum(axis=1)

    if divisors is not None:
        # print(f'Divisores para {new_col_name}: {divisors}')  # Línea de depuración
        for i, divisor in enumerate(divisors):
            df.loc[df.index[i], new_col_name] = df.loc[df.index[i], new_col_name] / divisor

    logging.info(f'Se aplican los divisores ({divisors}) a la columna.')

    if factor != 1.:
        df[new_col_name] *= factor

    if rounding is not None:
        df[new_col_name] = df[new_col_name].round(rounding)

    if position is not None:
        cols = df.columns.tolist()
        cols.insert(position, cols.pop(cols.index(new_col_name)))
        df = df.reindex(columns=cols)

    # print(f'Datos para {new_col_name}: {df[new_col_name]}')  # Línea de depuración
    logging.info(f'Datos para {new_col_name}: {df[new_col_name]}')

    return df
    subtract_cols = []
    cols_to_sum = [col for col in df.columns if any(code in col for code in codes)]
    logging.info(f'Columns to be summed for {new_col_name}: {cols_to_sum}') 

    if subtract_codes is not None:
        subtract_cols = [col for col in df.columns if any(code in col for code in subtract_codes)]
        for col in subtract_cols:
            if col in cols_to_sum:  # check if column is in cols_to_sum before removing
                cols_to_sum.remove(col)  

    logging.info(f'Columns to be subtracted for {new_col_name}: {subtract_cols}')
    # print(f'{new_col_name}: {cols_to_sum}')  # Debug line
    # print(f'Columns to be subtracted for {new_col_name}: {subtract_cols}')  # Debug line

    df[new_col_name] = df[cols_to_sum].sum(axis=1)
    df[new_col_name] -= df[subtract_cols].sum(axis=1)

    if divisors is not None:
        # print(f'Divisors for {new_col_name}: {divisors}')  # Debug line
        for i, divisor in enumerate(divisors):
            df.loc[df.index[i], new_col_name] = df.loc[df.index[i], new_col_name] / divisor

    logging.info(f'Divisors ({divisors}) are applied to the column.')

    if factor != 1.:
        df[new_col_name] *= factor

    if rounding is not None:
        df[new_col_name] = df[new_col_name].round(rounding)

    if position is not None:
        cols = df.columns.tolist()
        cols.insert(position, cols.pop(cols.index(new_col_name)))
        df = df.reindex(columns=cols)

    # print(f'Data for {new_col_name}: {df[new_col_name]}')  # Debug line
    logging.info(f'Data for {new_col_name}: {df[new_col_name]}')

    return df

def create_output_dataframe(data_dict, col_final):
    # Crear un DataFrame con las fechas recogidas como filas y las columnas de los códigos
    output_df = pd.DataFrame(data_dict).T

    # Cambiar el nombre de las columnas a los nombres finales
    output_df.columns = col_final

    # Eliminar todas las columnas que solo tienen un encabezado y luego no tienen datos
    output_df = output_df.dropna(axis=1, how='all')

    # Eliminar todas las columnas con la palabra "total" en el nombre
    output_df = output_df[output_df.columns.drop(list(output_df.filter(regex='Total')))]

    # Crear una lista de tuplas donde cada tupla es (nombre_columna_original, código)
    column_tuples = [(col, col.split(':')[0]) for col in output_df.columns.tolist()]

    # Ordenar esta lista en base al código completo
    column_tuples.sort(key=lambda x: int(x[1].split('-')[0] + x[1].split('-')[1]))

    # Extraer la lista ordenada de nombres de columnas originales
    sorted_columns = [col[0] for col in column_tuples]

    # Reorganizar las columnas en el dataframe
    output_df = output_df[sorted_columns]

    return output_df

def create_custom_columns(output_df, divisors, new_position, column_definitions):
    logging.info('Hasta este punto, las fechas han poblado el data_dict y el output_df ha sido creado.')
    logging.info('Las columnas han sido renombradas a los nombres finales.')
    logging.info('Las columnas que solo tienen un encabezado y no tienen datos han sido eliminadas.')
    logging.info('Las columnas con la palabra "total" en el nombre han sido eliminadas.')
    logging.info('Las columnas han sido ordenadas en base al código completo.')
    logging.info('Las columnas han sido reorganizadas en el dataframe.')
    for column_definition in column_definitions:
        if 'codes' in column_definition:
            # Funcionalidad existente
            output_df = create_and_format_column(
                df=output_df,
                new_col_name=column_definition['new_col_name'],
                codes=column_definition['codes'],
                divisors=divisors,
                position=new_position,
                factor=column_definition.get('factor', 1),
                rounding=column_definition.get('rounding', 0),
                subtract_codes=column_definition.get('subtract_codes', None)
            )
        elif 'sum_columns' in column_definition:
            # Nueva funcionalidad
            output_df[column_definition['new_col_name']] = output_df[column_definition['sum_columns']].sum(axis=1)
        elif 'divisor' in column_definition:
            # Crear columna a partir de los divisores
            output_df[column_definition['new_col_name']] = divisors 

    logging.info('Se han creado columnas personalizadas.')
    return output_df

def style_and_save_excel(temp_filename, final_filename):
    wb = openpyxl.load_workbook(temp_filename)
    ws = wb['Sheet1']

    # Establecer negrita para encabezados y color para encabezados de columnas
    header_font = Font(bold=True)
    header_fill = PatternFill(fill_type="solid", fgColor="9ab7e6")

    # Añadir franjas alternas a las filas
    stripe_fill_gray = PatternFill(fill_type="solid", fgColor="D3D3D3")  # Color gris claro
    stripe_fill_white = PatternFill(fill_type="solid", fgColor="FFFFFF")  # Color blanco

    # Establecer bordes de celda estándar
    thin_border = Border(
        left=Side(border_style="thin", color="d3d3d3"),
        right=Side(border_style="thin", color="d3d3d3"),
        top=Side(border_style="thin", color="d3d3d3"),
        bottom=Side(border_style="thin", color="d3d3d3"),
    )

    logging.info('Se han establecido estilos para bordes, franjas y encabezados')

    # Dict para almacenar la longitud máxima de cada columna
    col_max_length = {}

    for row in ws.iter_rows():
        for cell in row:
            i = cell.column_letter

            # Calcular la longitud máxima de la columna sin convertir int a str
            if isinstance(cell.value, int):
                cell_length = len(f"{cell.value}")
            else:
                cell_length = len(str(cell.value))
            
            if i not in col_max_length:
                col_max_length[i] = cell_length
            else:
                if cell_length > col_max_length[i]:
                    col_max_length[i] = cell_length

            # Aplicar estilos a los encabezados
            if cell.row == 1:
                cell.font = header_font
                cell.fill = header_fill
            else:
                # Aplicar franjas alternas a todas las demás filas
                fill = stripe_fill_gray if cell.row % 2 == 0 else stripe_fill_white
                cell.fill = fill

            # Aplicar bordes a todas las celdas
            cell.border = thin_border

            # Aplicar formato de contabilidad a todas las celdas numéricas
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0_);(#,##0)'

    logging.info('Se han aplicado estilos al archivo de Excel.') 

    # Ajustar la longitud de cada columna
    for i, length in col_max_length.items():
        ws.column_dimensions[i].width = length + 5

    # Renombrar 'Sheet1' a 'Salida Final'
    ws.title = 'Salida Final'

    # Guardar el archivo como UF_modelo_{nombre del archivo original}
    wb.save(f'UF_modelo_{os.path.basename(final_filename)}')

    logging.info(f'El archivo ha sido guardado como UF_modelo_{os.path.basename(final_filename)} en {os.getcwd()}')
    
    # Imprimir que el archivo ha sido procesado y guardado bajo el nombre UF_modelo_{nombre del archivo original} y mostrar el directorio donde está el nuevo archivo 
    print(f'El archivo ha sido procesado y guardado bajo el nombre UF_modelo_{os.path.basename(final_filename)} en {os.getcwd()}')
    wb = openpyxl.load_workbook(temp_filename)
    ws = wb['Sheet1']

def export_to_excel(output_df, filename):
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=True) as temp:
        output_df.T.to_excel(temp.name, index=True)
        style_and_save_excel(temp.name, filename)