# Importar todas las bibliotecas necesarias
import pandas as pd
import tkinter as tk
from tkinter import filedialog
import os
import requests
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side
import tempfile
import logging


def get_file_path():
    input('Presione Enter para seleccionar un archivo para leer en un dataframe...')
    root = tk.Tk()
    root.withdraw()
    file_path = filedialog.askopenfilename()

    # Asegurarse de que el archivo seleccionado sea un archivo de Excel que pueda ser procesado
    if not file_path.endswith('.xlsx'):
        print('Por favor, elija un archivo de Excel.')
        exit()

    # Crear una nueva carpeta llamada "UF_model_{nombre de archivo original}_{fecha actual}"
    new_folder_name = f'UF_model_{os.path.basename(file_path)}_{datetime.now().strftime("%Y-%d-%m_%H-%M-%S")}'
    if not os.path.exists(new_folder_name):
        os.mkdir(new_folder_name)

    # Cambiar el directorio de trabajo actual a la nueva carpeta
    os.chdir(new_folder_name)

    # Configuración del logging
    log_file_name = f'log_{os.path.basename(file_path)}_{datetime.now().strftime("%Y-%d-%m_%H-%M-%S")}.log'
    logging.basicConfig(filename=log_file_name, level=logging.INFO, 
                        format='%(asctime)s %(levelname)s %(message)s', 
                        datefmt='%m/%d/%Y %I:%M:%S %p',
                        )
    
    logging.info('Inició la función get_file_path')

    # Mensaje de diálogo confirmando que el archivo ha sido seleccionado y leído
    print(f'El archivo {os.path.basename(file_path)} ha sido leído en un dataframe.')
    logging.info(f'El archivo {os.path.basename(file_path)} ha sido leído en un dataframe.')

    return file_path

def load_dataframe(file_path):
    logging.info(f'Cargando dataframe desde {file_path}')
    return pd.read_excel(file_path)

def find_code_column(df):
    code_column = None
    code_prefixes = [(4, 'REV'), (5, 'RE'), (6, 'NRE'), (7, 'C&GE'), (8, 'OI&E'), (9, 'OI&E')]
    groups = {i: {"codes": [], "names": [], "final": [], "prefix": prefix} for i, prefix in code_prefixes}
    logging.info(f'Se han creado los prefijos de código basados en el primer dígito del código: {code_prefixes}.')
    logging.info(f'Se han creado grupos basados en los prefijos de código: {groups}.')

    # Encontrar la columna que contiene los códigos basándose en el código de ingresos
    for col in df.columns:
        if '40100-00' in df[col].values:
            code_column = col
            break
    logging.info('Se ha encontrado la columna de código.')
    return code_column, code_prefixes, groups

def group_data_by_code(df, code_column, code_prefixes, groups):
    if code_column is None:
        print("No se pudo encontrar la columna de ingresos.")
        logging.info('No se pudo encontrar la columna de ingresos.')
    else:
        name_column = df.columns[df.columns.get_loc(code_column) + 1]
        logging.info(f'Se ha creado la columna de nombres basada en la posición de la columna de códigos.')

        # Para cada grupo, filtrar las filas basándose en el código, crear un diccionario y poblar las listas
        for i, prefix in code_prefixes:
            group_rows = df[(df[code_column] >= f'{i}0000-00') & (df[code_column] <= f'{i}9999-99')]
            group_data = {(code, f"{prefix} - " + group_rows.loc[group_rows[code_column] == code, name_column].values[0].strip().title()): [] for code in group_rows[code_column]}      
            for (code, name), data in group_data.items():
                groups[i]["codes"].append(code)
                groups[i]["names"].append(name)
                groups[i]["final"].append(code + ": " + name)

        logging.info('Se han poblado los grupos para el grupo basado en el prefijo del código.')
        logging.info('Se han creado las filas de grupo para el grupo basado en el prefijo del código.')
        logging.info('Se han creado los datos del grupo para el grupo basado en el prefijo del código.')

    return groups

def populate_data_dict(df, code_column, codes, dates):
    data_dict = {}
    # Iterate over each column
    for col in df.columns:
        # Iterate over each row in the column
        for cell in df[col]:
            try:
                # Try to convert the cell to a datetime
                date = pd.to_datetime(cell, format='%b %Y', errors='coerce')

                # If the cell is a date
                if pd.notnull(date):
                    date_str = cell
                    dates.append(date_str)
                    data_dict[date_str] = {}
                    

                    for code in codes:
                        # Get the row where the code matches in the original DataFrame
                        row = df[df[code_column] == code]

                        # If there's any matching row, get the corresponding value from the row with code
                        if not row.empty:
                            value = row.iat[0, df.columns.get_loc(col)]
                            data_dict[date_str][code] = value

                    # Exit inner loop once the first date is found in the column
                    break
            except ValueError:
                # If the cell couldn't be converted to a date, continue to the next cell
                continue
    return data_dict, dates

def fetch_exchange_rate(dates):
    try:
        # if server isn't down
        url = 'https://www.alphavantage.co/query?function=FX_MONTHLY&from_symbol=CLF&to_symbol=CLP&apikey=5TLIDN7TN6IQZJUE'
        r = requests.get(url)
        data = r.json()

        # If there is an error message in the response, return default divisors
        if 'Error Message' in data:
            print("Server is down. Using default divisors.")
            logging.info('Server is down. Using default divisors.')
            return [1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1]

        # Extract the monthly exchange rate data from the response
        monthly_rates = data['Time Series FX (Monthly)']

        # List to store the exchange rates
        exchange_rates = []

        # Iterate over the monthly data and extract the exchange rates
        for date, rate in monthly_rates.items():
            # Convert the date string to a datetime object
            date_obj = datetime.strptime(date, '%Y-%m-%d')
            # Format the date as "Month Year" (e.g., "Jun 2022")
            month_year = date_obj.strftime('%b %Y')

            # Check if the month/year is in the dates list
            if month_year in dates:
                exchange_rate = rate['4. close']
                exchange_rates.append((month_year, exchange_rate))

        # Sort the exchange rates based on the month/year in ascending order
        exchange_rates.sort(key=lambda x: datetime.strptime(x[0], '%b %Y'))

        # Extract the exchange rates from the list of tuples
        divisors_UF = [x[1] for x in exchange_rates]

        # Convert the exchange rates to floats
        divisors_UF = [float(x) for x in divisors_UF]

        print("Successfully fetched data from the server.")
        logging.info('Successfully fetched data from the server.')
        return divisors_UF

    except Exception as e:
        print(f'Error fetching exchange rate: {e}')
        print("Using default divisors due to the error.")
        logging.info(f'Error fetching exchange rate: {e}')
        logging.info('Using default divisors due to the error.')
        # In case of any other unhandled exception, return default divisors
        return [1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1]

def create_and_format_column(df, new_col_name, codes, divisors=None, rounding=0, position=None, factor=1., subtract_codes=None):
    subtract_cols = []
    cols_to_sum = [col for col in df.columns if any(code in col for code in codes)]
    logging.info(f'Columns to be summed for {new_col_name}: {cols_to_sum}') 

    if subtract_codes is not None:
        subtract_cols = [col for col in df.columns if any(code in col for code in subtract_codes)]
        for col in subtract_cols:
            if col in cols_to_sum:  # check if column is in cols_to_sum before removing
                cols_to_sum.remove(col)  

    logging.info(f'Columns to be subtracted for {new_col_name}: {subtract_cols}')
    # print(f'{new_col_name}: {cols_to_sum}')  # Debug line
    # print(f'Columns to be subtracted for {new_col_name}: {subtract_cols}')  # Debug line

    df[new_col_name] = df[cols_to_sum].sum(axis=1)
    df[new_col_name] -= df[subtract_cols].sum(axis=1)

    if divisors is not None:
        # print(f'Divisors for {new_col_name}: {divisors}')  # Debug line
        for i, divisor in enumerate(divisors):
            df.loc[df.index[i], new_col_name] = df.loc[df.index[i], new_col_name] / divisor

    logging.info(f'Divisors ({divisors}) are applied to the column.')

    if factor != 1.:
        df[new_col_name] *= factor

    if rounding is not None:
        df[new_col_name] = df[new_col_name].round(rounding)

    if position is not None:
        cols = df.columns.tolist()
        cols.insert(position, cols.pop(cols.index(new_col_name)))
        df = df.reindex(columns=cols)

    # print(f'Data for {new_col_name}: {df[new_col_name]}')  # Debug line
    logging.info(f'Data for {new_col_name}: {df[new_col_name]}')

    return df

def create_output_dataframe(data_dict, col_final):
    # Create a DataFrame with the collected dates as rows and the columns from codes
    output_df = pd.DataFrame(data_dict).T

    # Rename the columns to the final names
    output_df.columns = col_final

    # delete all columns that do not have only have a header and then no data
    output_df = output_df.dropna(axis=1, how='all')

    # delete all columns with the word "total" in the name
    output_df = output_df[output_df.columns.drop(list(output_df.filter(regex='Total')))]

    # Create a list of tuples where each tuple is (original_column_name, code)
    column_tuples = [(col, col.split(':')[0]) for col in output_df.columns.tolist()]

    # Sort this list based on the entire code
    column_tuples.sort(key=lambda x: int(x[1].split('-')[0] + x[1].split('-')[1]))

    # Extract the sorted list of original column names
    sorted_columns = [col[0] for col in column_tuples]

    # Rearrange the columns in the dataframe
    output_df = output_df[sorted_columns]

    return output_df

def create_custom_columns(output_df, divisors, new_position, column_definitions):
    logging.info('By this point, dates have populated the data_dict and the output_df has been created.')
    logging.info('Columns have been renamed to the final names.')
    logging.info('Columns with only a header and no data have been deleted.')
    logging.info('Columns with the word "total" in the name have been deleted.')
    logging.info('Columns have been sorted based on the entire code.')
    logging.info('Columns have been rearranged in the dataframe.')
    for column_definition in column_definitions:
        if 'codes' in column_definition:
            # Existing functionality
            output_df = create_and_format_column(
                df=output_df,
                new_col_name=column_definition['new_col_name'],
                codes=column_definition['codes'],
                divisors=divisors,
                position=new_position,
                factor=column_definition.get('factor', 1),
                rounding=column_definition.get('rounding', 0),
                subtract_codes=column_definition.get('subtract_codes', None)
            )
        elif 'sum_columns' in column_definition:
            # New functionality
            output_df[column_definition['new_col_name']] = output_df[column_definition['sum_columns']].sum(axis=1)
        elif 'divisor' in column_definition:
            # Create column from the divisors
            output_df[column_definition['new_col_name']] = divisors 

    logging.info(f'Custom columns have been created.')
    return output_df

def style_and_save_excel(temp_filename, final_filename):
    wb = openpyxl.load_workbook(temp_filename)
    ws = wb['Sheet1']

    # Set font bold for headers and color for column headers
    header_font = Font(bold=True)
    header_fill = PatternFill(fill_type="solid", fgColor="9ab7e6")

    # Add alternating stripes to rows
    stripe_fill_gray = PatternFill(fill_type="solid", fgColor="D3D3D3")  # Light Gray color
    stripe_fill_white = PatternFill(fill_type="solid", fgColor="FFFFFF")  # White color

    # Set standard cell borders
    thin_border = Border(
        left=Side(border_style="thin", color="d3d3d3"),
        right=Side(border_style="thin", color="d3d3d3"),
        top=Side(border_style="thin", color="d3d3d3"),
        bottom=Side(border_style="thin", color="d3d3d3"),
    )

    logging.info(f'Styles have been set for border, stripes, and headers)')

    # Dict to store max length of each column
    col_max_length = {}

    for row in ws.iter_rows():
        for cell in row:
            i = cell.column_letter

            # Calculate max column length without converting int to str
            if isinstance(cell.value, int):
                cell_length = len(f"{cell.value}")
            else:
                cell_length = len(str(cell.value))
            
            if i not in col_max_length:
                col_max_length[i] = cell_length
            else:
                if cell_length > col_max_length[i]:
                    col_max_length[i] = cell_length

            # Apply styles to headers
            if cell.row == 1:
                cell.font = header_font
                cell.fill = header_fill
            else:
                # Apply alternating stripes to all other rows
                fill = stripe_fill_gray if cell.row % 2 == 0 else stripe_fill_white
                cell.fill = fill

            # Apply borders to all cells
            cell.border = thin_border

            # Apply Accounting format to all numeric cells
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0_);(#,##0)'

    logging.info(f'Styles have been applied to the excel file.') 

    # Adjusting the length of each column
    for i, length in col_max_length.items():
        ws.column_dimensions[i].width = length + 5

    # rename 'Sheet1' to 'Final Output'
    ws.title = 'Final Output'

    # save the file as UF_model_{original file name}
    wb.save(f'UF_model_{os.path.basename(final_filename)}')

    logging.info(f'File has been saved as UF_model_{os.path.basename(final_filename)} in {os.getcwd()}')
    # print that the file has been processed and saved under the name processed_model_{original file name} and show the directory where the new file is 
    print(f'File has been processed and saved under the name UF_model_{os.path.basename(final_filename)} in {os.getcwd()}')

def export_to_excel(output_df, filename):
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=True) as temp:
        output_df.T.to_excel(temp.name, index=True)
        style_and_save_excel(temp.name, filename)